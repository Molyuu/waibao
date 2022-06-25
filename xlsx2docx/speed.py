from docxtpl import DocxTemplate
from openpyxl import Workbook, load_workbook
from docx import Document

def find_excel(xlsx):
    #读取Excel中的第一行内容
    cell=[]
    wb=load_workbook(filename=xlsx)
    sheet=wb.active
    for i in sheet[1]:
        cell.append(i.value)
    return cell

def find_word(docx,word):
    document = Document(docx)
    all_paragraphs = document.paragraphs
    list1 = []
    for paragraph in all_paragraphs:
 
        str1 = paragraph.text
        if str1.find(word) != -1 :
            return True
        else:
            return False
    

def replace_word(docx,xlsx):
    doc1=DocxTemplate(docx) 
    list_row=[]
    for i in find_excel(xlsx):
        if find_word(docx,i):
            pass
            


'''
def find(xlsx):
    #读取Excel中的第一行内容
    cell=[]
    wb=load_workbook(filename=xlsx)
    sheet=wb.active
    for i in sheet[1]:
        cell.append(i.value)
    return cell

def replace(xlsx,docx):
    doc=DocxTemplate(docx)
    wb=load_workbook(filename=xlsx)
    sheet=wb.active
    for j in find(xlsx):
        for k in sheet:
            context= {'{%s}'% j : ""}
    
'''
if __name__ == "__main__":
    pass