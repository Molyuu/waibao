from openpyxl import load_workbook


def column_to_name(colnum):
    str = ""
    while not (colnum // 26 == 0 and colnum % 26 == 0):
        temp = 25
        if colnum % 26 == 0:
            str += chr(temp + 65)
        else:
            str += chr(colnum % 26 - 1 + 65)
        colnum //= 26
    return str


wb = load_workbook("./new.xlsx", data_only=True)
sheet = wb.active

for i in range(sheet.max_row):
    for j in range(sheet.max_column):
        """
        print(i + 1, column_to_name(j + 1))
        """
        cell_type = sheet["%s%d" % (column_to_name(j + 1), i + 1)].number_format
        print(cell_type)
